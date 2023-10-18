from selenium import webdriver
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import csv
from selenium.webdriver.common.by import By

# Instantiate the Selenium driver (replace with the appropriate webdriver for your browser)
driver = webdriver.Chrome()

# Navigate to the desired webpage
driver.get("https://online.kfc.co.in/menu/burgers")  # Replace with the actual webpage URL
sleep(2)

# Find the element using a suitable locator
elements= driver.find_elements(By.CLASS_NAME, 'medium-menu-product-card')

product_list = []

# Iterate over the found elements
for element in elements:
# Extract the required information
    img_url = element.find_element(By.TAG_NAME, 'img').get_attribute('src')
    product_name = element.find_element(By.CLASS_NAME, 'medium-menu-product-header').text
    price = element.find_element(By.CLASS_NAME, 'medium-menu-product-price').text
    description = element.find_element(By.ID, 'longDescription').text

# Print the extracted information
    product_list.append({
        "Image Source": img_url,
        "Product Name": product_name,
        "Price": price,
        "Description": description
    })
print("prod list ,\n", product_list)

# Save the data to an Excel file
excel_file = "products.xlsx"
excel_columns = ["Image Source", "Product Name", "Price", "Description"]

try:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the column names
    for col_num, column_title in enumerate(excel_columns, start=1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # Write the data rows
    for row_num, data in enumerate(product_list, start=2):
        for col_num, column_title in enumerate(excel_columns, start=1):
            sheet.cell(row=row_num, column=col_num, value=data[column_title])

    # Save the workbook
    workbook.save(filename=excel_file)
    print("Data saved to", excel_file)
except IOError:
    print("I/O error")

# Close the browser
driver.quit()

# Close the browser
driver.quit()
